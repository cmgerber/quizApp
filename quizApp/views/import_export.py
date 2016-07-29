"""Functions for importing and exporting data via XLSX files.

When exporting data, we grab everything from the database and exclude the
columns that have export_include: False in their info attributes.

When generating a template for imports, we include exclude columns that have
import_include: False in their info attributes.

When importing from a spreadsheet, we will process every column (regardless of
info attribute). However, in some cases certain columns need to be filled out
earlier than others. This is why we use a special method on each model to
populate an object.
"""
import os
from collections import OrderedDict, defaultdict
import tempfile

from openpyxl import Workbook
from sqlalchemy import inspect
from sqlalchemy.orm.interfaces import ONETOMANY, MANYTOMANY
from sqlalchemy.orm.properties import ColumnProperty, RelationshipProperty

from quizApp import db
from quizApp import models

SHEET_NAME_MAPPING = OrderedDict([
    ("Experiments", models.Experiment),
    ("Participant Experiments", models.ParticipantExperiment),
    ("Datasets", models.Dataset),
    ("Media items", models.MediaItem),
    ("Assignments", models.Assignment),
    ("Activities", models.Activity),
    ("Choices", models.Choice),
])


def export_to_workbook():
    """Retrieve elements from thed database and save them to a workbook.
    Return the filename of the workbook.

    When we export to a workbook, we export almost everything. To prevent
    a field from being exported, add `"export_include": False` to a field's
    `info` dictionary.
    """

    workbook = Workbook()
    workbook.remove_sheet(workbook.active)

    for sheet_name, model in SHEET_NAME_MAPPING.iteritems():
        current_sheet = workbook.create_sheet()
        current_sheet.title = sheet_name
        sheet_data = object_list_to_sheet(model.query.all())
        write_list_to_sheet(sheet_data, current_sheet)

    file_name = tempfile.mkstemp(".xlsx")
    os.close(file_name[0])
    workbook.save(file_name[1])
    return file_name[1]


def include_column(model, column, prop, key):
    """Determine if a column should be included in exporting.
    """
    include = True
    if isinstance(prop, ColumnProperty):
        info = prop.columns[0].info
    elif isinstance(prop, RelationshipProperty):
        info = prop.info

    include = info.get(key, True)

    # If we are looking at a foreign key and this foreign key also
    # includes a backref field, don't put the foreign key into the sheet
    # it is better to include the backref because we automatically run
    # validation on the backref
    if column[-3:] == "_id" and \
            column[:-3] in inspect(model).attrs and \
            key not in info:
        include = False

    return include


def header_from_property(prop):
    """Given a property, return its name for use in sheet headers.
    """
    return prop.parent.class_.__tablename__ + "_" + prop.key


def header_to_field_name(header, model):
    """Reverse header_from_property - given a header and a model, return the
    actual name of the field.
    """
    prefix = model.__tablename__ + "_"

    if header[:len(prefix)] != prefix:
        # Sanity check failed
        raise ValueError("Incorrect header/model combination")

    return header[len(prefix):]


def model_to_sheet_headers(model):
    """Given a model, put all columns whose info contains "import_include":
    True into a list of headers.

    This method also inspects all polymorphisms of the given model.
    """
    headers = []

    polymorphisms = inspect(model).mapper.polymorphic_map.values()
    if not polymorphisms:
        polymorphisms = [model]

    seen_columns = set()

    for polymorphism in polymorphisms:
        for column, prop in get_field_order(polymorphism):
            if include_column(model, column, prop, "import_include") and \
                    column not in seen_columns:
                headers.append(header_from_property(prop))
                seen_columns.add(column)

    return [headers]


def write_list_to_sheet(data_list, sheet):
    """Given a list and a sheet, write out the list to the sheet.
    The list should be two dimensional, with each list in the list representing
    a row.
    """

    for r in xrange(1, len(data_list) + 1):
        for c in xrange(1, len(data_list[r - 1]) + 1):
            sheet.cell(row=r, column=c).value = data_list[r - 1][c - 1]


def relationship_to_string(field, value):
    """Given a relationship, convert it to a comma separated list of integers.
    """
    ids = []
    direction = field.property.direction
    if direction in (MANYTOMANY, ONETOMANY):
        for obj in value:
            ids.append(str(obj.id))
        return ",".join(ids)
    elif value:
        return str(value.id)
    return ""


def field_to_string(obj, column):
    """Given an object and a column name, convert the contents
    of the column to a string. If it is a collection, return a
    comma separated list of ids.
    """
    model = type(obj)
    value = getattr(obj, column)
    field_attrs = inspect(model).attrs[column]

    if isinstance(field_attrs, RelationshipProperty):
        value = relationship_to_string(getattr(model, column), value)

    return value


def get_field_order(model):
    """Given a model, return a list of properties in the order given by that
    model's Meta class.
    """
    try:
        field_order = model.Meta.field_order
    except AttributeError:
        field_order = ('*',)

    visited = set()
    ordered_fields = []
    fields = {f[0]: f for f in inspect(model).attrs.items()}

    # Reorder the fields based on field_order, if it exists.
    for field_mask in field_order:
        if field_mask not in visited:
            if field_mask == '*':
                for field_name, field in fields.items():
                    if field_name in visited or field_name in field_order:
                        continue
                    ordered_fields.append(field)
            elif field_mask in fields:
                ordered_fields.append(fields[field_mask])
            visited.add(field_mask)

    return ordered_fields


def object_list_to_sheet(object_list):
    """Given a list of objects, iterate over all of them and create a list
    of lists that can be written using write_list_to_sheet.

    The first row returned will be a header row. All fields will be included
    unless a field contains export_include: False in its info attibute.
    """
    sheet = [[]]
    for obj in object_list:
        row = [""] * len(sheet[0])

        ordered_fields = get_field_order(type(obj))

        for column, prop in ordered_fields:
            include = include_column(type(obj), column, prop, "export_include")

            if not include:
                continue

            if column not in sheet[0]:
                sheet[0].append(column)
                row.append("")

            index = sheet[0].index(column)

            row[index] = field_to_string(obj, column)
        sheet.append(row)
    return sheet


def populate_field(obj, field_name, value, pk_mapping, args):
    """Populate a field on a certain object based on the value from an imported
    spreadsheet. The field on obj will not be set, rather args will be
    populated with the field's name and its value for use with the model's
    import_dict method.

    This may involve doing a database lookup if the field in question is a
    relationship field.

    A note on pk_mapping:

    To avoid conflicts between imported PK and existing PK, we do not
    assign PK's based on user input. However we have to store them
    because other rows in the user input may be referencing a certain PK. So
    we store them in a kind of bastard mini-database in memory while we
    are importing data.

    Arguments:
        obj - The object whose fields need populating.
        field_name - A string containing the name of the field that should be
        populated.
        value - The value of the field, as read from the spreadsheet.
        pk_mapping - A mapping of any objects created in this import session
        that value may refer to.
        args - A dictionary to set the field and its value in.
    """
    model = type(obj)
    field_attrs = inspect(model).attrs[field_name]
    field = getattr(model, field_name)
    if isinstance(field_attrs, RelationshipProperty):
        # This is a relationship
        remote_model = field.property.mapper.class_
        direction = field.property.direction
        if direction in (MANYTOMANY, ONETOMANY):
            args[field_name] = []
            values = str(value).split(",")
            for fk in values:
                fk = int(float(fk))  # goddamn stupid excel
                collection_item = get_object_from_id(remote_model, fk,
                                                     pk_mapping)
                if collection_item:
                    args[field_name].append(collection_item)
        else:
            value = int(float(value))  # goddamn stupid excel
            collection_item = get_object_from_id(remote_model, value,
                                                 pk_mapping)
            if collection_item:
                args[field_name] = collection_item
    elif field.primary_key:
        pk_mapping[model.__tablename__][int(float(value))] = obj
    elif isinstance(field_attrs, ColumnProperty):
        args[field_name] = value


def get_object_from_id(model, obj_id, pk_mapping):
    """If the object of type model and id obj_id is in pk_mapping,
    return it. Otherwise, query the database.
    """
    try:
        return pk_mapping[model.__tablename__][obj_id]
    except KeyError:
        return model.query.get(obj_id)


def instantiate_model(model, headers, row):
    """Return an instance of a model, correctly handling polymorphism.

    Since the actual object represented by a row may be different from another
    in the same sheet, we need to take into account any possible polymorphisms.

    This method looks at the model to see if it has polymorphic identities. If
    so, it returns an instance of the correct one. If not, it returns an
    instance of the passed in model.
    """
    model_mapper = inspect(model).mapper
    if not model_mapper.polymorphic_identity:
        return model()

    polymorphic_index = headers.index(model_mapper.polymorphic_on.name)
    polymorphic_type = row[polymorphic_index].value
    return model_mapper.polymorphic_map[polymorphic_type].class_()


def import_data_from_workbook(workbook):
    """Given an excel workbook, read in the sheets and save them to the
    database.
    """
    pk_mapping = defaultdict(dict)

    for sheet_name, model in SHEET_NAME_MAPPING.iteritems():
        try:
            sheet = workbook.get_sheet_by_name(sheet_name)
        except KeyError:
            # This model is not present in the workbook
            continue

        headers = [header_to_field_name(c.value, model) for c in sheet.rows[0]]

        for row in sheet.rows[1:]:
            obj = instantiate_model(model, headers, row)
            obj_args = {}

            for col_index, cell in enumerate(row):
                if not cell.value:
                    continue
                populate_field(obj, headers[col_index],
                               cell.value, pk_mapping, obj_args)

            if obj_args:
                obj.import_dict(**obj_args)
                db.session.add(obj)

        db.session.commit()
