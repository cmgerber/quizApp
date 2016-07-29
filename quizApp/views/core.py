"""This blueprint takes care of rendering static pages outside of the other
blueprints.
"""
import os
from collections import OrderedDict
import tempfile

import openpyxl
from flask import Blueprint, render_template, send_file, jsonify
from flask_security import roles_required

from quizApp import models
from quizApp.views import import_export
from quizApp.forms.core import ImportDataForm


core = Blueprint("core", __name__, url_prefix="/")


# homepage
@core.route('')
def home():
    """Display the homepage."""
    return render_template('core/index.html',
                           is_home=True)


@core.route('export')
@roles_required("experimenter")
def export_data():
    """Send the user a breakddown of datasets, activities, etc. for use in
    making assignments.
    """
    file_name = import_export.export_to_workbook()
    return send_file(file_name, as_attachment=True,
                     attachment_filename="quizapp_export.xlsx")


@core.route('import_template')
@roles_required("experimenter")
def import_template():
    """Send the user a blank excel sheet that can be filled out and used to
    populate an experiment's activity list.

    The process is essentially:

    1. Get a list of models to include

    2. From each model, get all its polymorphisms

    3. For each model, get all fields that should be included in the import
    template, including any fields from polymorphisms

    4. Create a workbook with as many sheets as models, with one row in each
    sheet, containing the name of the included fields
    """

    sheets = OrderedDict([
        ("Experiments", models.Experiment),
        ("Participant Experiments", models.ParticipantExperiment),
        ("Assignments", models.Assignment),
        ("Activities", models.Activity),
        ("Choices", models.Choice),
    ])

    documentation = [
        ["Do not modify the first row in every sheet!"],
        ["Simply add in your data in the rows undeneath it."],
        ["Use IDs from the export sheet to populate relationship columns."],
        [("If you want multiple objects in a relation, separate the IDs using"
          " commas.")],
        [("There is no need to modify any sheet if you are not interested in "
          "adding in objects of that type")],
        [("Example: To make an experiment and use existing assignments for "
          "the experiment, fill out the Experiments, Participant Experiment, "
          "and Assignment sheets.")],
        [("Example: To add assignments to an existing experiment, fill out "
          "the Participant Experiment and Assignment sheet. For the "
          "participant_experiment_experiment column, use the experiment_id "
          "of the experiment you wish to modify. You can find this ID in the "
          "export spreadsheet.")],
        [("If you wish to do one the above as well as create new "
          "activities, fill out the sheets mentioned above as well as the "
          "Activities sheet. If you are making new multiple choice questions, "
          "you'll also need to fill out the Choices sheet.")],
    ]

    workbook = openpyxl.Workbook()
    workbook.remove_sheet(workbook.active)

    current_sheet = workbook.create_sheet()
    current_sheet.title = "Documentation"
    import_export.write_list_to_sheet(documentation, current_sheet)

    for sheet_name, model in sheets.iteritems():
        current_sheet = workbook.create_sheet()
        current_sheet.title = sheet_name
        headers = import_export.model_to_sheet_headers(model)
        import_export.write_list_to_sheet(headers, current_sheet)


    file_handle, file_name = tempfile.mkstemp(".xlsx")
    os.close(file_handle)
    workbook.save(file_name)
    return send_file(file_name, as_attachment=True,
                     attachment_filename="import_template.xlsx")


@core.route('import', methods=["POST"])
@roles_required("experimenter")
def import_data():
    """Given an uploaded spreadsheet, import data from the spreadsheet
    into the database.
    """
    import_data_form = ImportDataForm()

    if not import_data_form.validate():
        return jsonify({"success": 0, "errors": import_data_form.errors})

    workbook = openpyxl.load_workbook(import_data_form.data.data)

    import_export.import_data_from_workbook(workbook)

    return jsonify({"success": 1})


@core.route('manage_data', methods=["GET"])
@roles_required("experimenter")
def manage_data():
    """Show a form for uploading data and such.
    """

    import_data_form = ImportDataForm()

    return render_template("core/manage_data.html",
                           import_data_form=import_data_form)
